"""
Excel formatting utilities for worksheets.
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .constants import (
    ROW_NUMBER_COLUMN_NAME,
    EXCEL_HEADER_BG_COLOR,
    EXCEL_HEADER_FONT_COLOR,
    EXCEL_BORDER_COLOR,
    EXCEL_HEADER_FONT_SIZE,
    EXCEL_MAX_COLUMN_WIDTH,
    EXCEL_COLUMN_WIDTH_PADDING,
    EXCEL_HEADER_ROW_HEIGHT,
    EXCEL_DATETIME_FORMAT
)


def add_row_numbers(df: pd.DataFrame) -> pd.DataFrame:
    """
    Add sequential row numbers starting from 1 at the beginning of DataFrame.
    """
    df.insert(0, ROW_NUMBER_COLUMN_NAME, range(1, len(df) + 1))
    return df


def format_worksheet(worksheet: Worksheet, df: pd.DataFrame) -> None:
    """
    Format worksheet for better readability and aesthetics.
    - Auto-adjust column widths
    - Format headers (bold, background color)
    - Add borders
    - Freeze top row
    """
    # Define styles
    header_font = Font(bold=True, color=EXCEL_HEADER_FONT_COLOR, size=EXCEL_HEADER_FONT_SIZE)
    header_fill = PatternFill(start_color=EXCEL_HEADER_BG_COLOR, end_color=EXCEL_HEADER_BG_COLOR, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    cell_alignment = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color=EXCEL_BORDER_COLOR),
        right=Side(style='thin', color=EXCEL_BORDER_COLOR),
        top=Side(style='thin', color=EXCEL_BORDER_COLOR),
        bottom=Side(style='thin', color=EXCEL_BORDER_COLOR)
    )
    
    # Format header row
    for col_num, column in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Format data cells and calculate column widths
    for col_num, column in enumerate(df.columns, 1):
        column_letter = get_column_letter(col_num)
        
        # Calculate maximum width for this column
        max_length = len(str(column))  # Start with header length
        
        # Check if this is a datetime column
        is_datetime_col = pd.api.types.is_datetime64_any_dtype(df[column])
        
        for row_num, value in enumerate(df[column], 2):  # Start from row 2 (data rows)
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.alignment = cell_alignment
            cell.border = thin_border
            
            # Format datetime cells
            if is_datetime_col and pd.notna(value):
                cell.number_format = EXCEL_DATETIME_FORMAT
            
            # Update max length
            if value is not None:
                cell_length = len(str(value))
                max_length = max(max_length, cell_length)
        
        # Set column width (add some padding)
        adjusted_width = min(max_length + EXCEL_COLUMN_WIDTH_PADDING, EXCEL_MAX_COLUMN_WIDTH)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze top row (header)
    worksheet.freeze_panes = worksheet['A2']
    
    # Set row height for header
    worksheet.row_dimensions[1].height = EXCEL_HEADER_ROW_HEIGHT
