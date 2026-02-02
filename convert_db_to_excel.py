#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Simple and reliable SQLite to Excel converter.
All tables from the database are exported as separate sheets to a single Excel file.
"""

import sqlite3
import os
import glob
import pandas as pd
import logging
import logging.config
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Setup logging
def setup_logging() -> logging.Logger:
    """Configure logging from configuration file"""
    log_dir = Path('logs')
    log_dir.mkdir(exist_ok=True)
    
    config_path = Path('config/logging.ini')
    if config_path.exists():
        logging.config.fileConfig(config_path)
    else:
        # Fallback to basic config if file not found
        logging.basicConfig(
            level=logging.INFO,
            format='%(message)s'
        )
    return logging.getLogger(__name__)


def find_all_db_files(input_dir: str = 'input') -> list[str]:
    """Find all .db files in the input/ folder"""
    db_files = glob.glob(os.path.join(input_dir, '*.db'))
    
    if not db_files:
        raise FileNotFoundError(f"No .db files found in '{input_dir}/' folder")
    
    return db_files


def get_output_path(db_path: str, output_dir: str = 'output') -> str:
    """Generate output Excel file path based on input database filename"""
    # Get the base filename without extension
    base_name = os.path.splitext(os.path.basename(db_path))[0]
    # Create output path with .xlsx extension
    output_path = os.path.join(output_dir, f"{base_name}.xlsx")
    return output_path


def get_all_tables(db_path: str) -> list[str]:
    """Get a list of all tables from the SQLite database"""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Get all tables (excluding SQLite system tables)
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'")
    tables = [row[0] for row in cursor.fetchall()]
    
    conn.close()
    return tables


def is_unix_timestamp_column(series: pd.Series) -> bool:
    """
    Check if a column contains Unix timestamps.
    Returns True if column appears to contain Unix timestamps.
    """
    # Check if column name suggests timestamp
    name_lower = str(series.name).lower()
    if not any(keyword in name_lower for keyword in ['time', 'timestamp', 'date']):
        return False
    
    # Check if values are numeric and in reasonable Unix timestamp range
    if not pd.api.types.is_numeric_dtype(series):
        return False
    
    # Unix timestamp range: 1/1/2000 (946684800) to 1/1/2100 (4102444800)
    non_null = series.dropna()
    if len(non_null) == 0:
        return False
    
    return non_null.between(946684800, 4102444800).all()


def convert_timestamps_to_readable(df: pd.DataFrame) -> pd.DataFrame:
    """
    Detect and convert Unix timestamp columns to readable datetime format.
    Adds new columns with '_readable' suffix next to original timestamp columns.
    """
    new_columns = {}
    insert_positions = {}
    
    for col in df.columns:
        if is_unix_timestamp_column(df[col]):
            # Create readable column name
            readable_col = f"{col}_readable"
            
            # Convert Unix timestamp to datetime
            new_columns[readable_col] = pd.to_datetime(df[col], unit='s', errors='coerce')
            
            # Store position to insert after original column
            insert_positions[readable_col] = list(df.columns).index(col) + 1
    
    # Insert new columns at appropriate positions
    for col_name in sorted(insert_positions.keys(), key=lambda x: insert_positions[x], reverse=True):
        position = insert_positions[col_name]
        df.insert(position, col_name, new_columns[col_name])
    
    return df


def rename_data_format_columns(df: pd.DataFrame, db_path: str) -> pd.DataFrame:
    """
    Rename data_format_X columns to readable names from data_format table.
    Maps data_format_0, data_format_1, etc. to their comment descriptions.
    """
    try:
        with sqlite3.connect(db_path) as conn:
            # Read the data_format table to get column descriptions
            format_df = pd.read_sql_query("SELECT data_format_index, comment FROM data_format", conn)
            
            # Create mapping from data_format_X to comment
            rename_map = {}
            for _, row in format_df.iterrows():
                col_name = f"data_format_{row['data_format_index']}"
                if col_name in df.columns:
                    rename_map[col_name] = row['comment']
            
            # Rename columns
            if rename_map:
                df = df.rename(columns=rename_map)
        
    except Exception:
        # If data_format table doesn't exist or has different structure, skip renaming
        pass
    
    return df


def add_row_numbers(df: pd.DataFrame) -> pd.DataFrame:
    """
    Add sequential row numbers starting from 1 at the beginning of DataFrame.
    """
    df.insert(0, 'Row', range(1, len(df) + 1))
    return df


def format_worksheet(worksheet: Worksheet, df: pd.DataFrame) -> None:
    """
    Format worksheet for better readability and aesthetics.
    - Auto-adjust column widths
    - Format headers (bold, background color)
    - Add borders
    - Freeze top row
    """
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    cell_alignment = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # Format header row
    for col_num, column in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Format data cells and calculate column widths
    for col_num, column in enumerate(df.columns, 1):
        column_letter = get_column_letter(col_num)
        
        # Calculate maximum width for this column
        max_length = len(str(column))  # Start with header length
        
        # Check if this is a datetime column
        is_datetime_col = pd.api.types.is_datetime64_any_dtype(df[column])
        
        for row_num, value in enumerate(df[column], 2):  # Start from row 2 (data rows)
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.alignment = cell_alignment
            cell.border = thin_border
            
            # Format datetime cells
            if is_datetime_col and pd.notna(value):
                cell.number_format = 'YYYY-MM-DD HH:MM:SS'
            
            # Update max length
            if value is not None:
                cell_length = len(str(value))
                max_length = max(max_length, cell_length)
        
        # Set column width (add some padding)
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze top row (header)
    worksheet.freeze_panes = worksheet['A2']
    
    # Set row height for header
    worksheet.row_dimensions[1].height = 20


def convert_db_to_excel(db_path: str, output_path: str = 'output/database.xlsx', logger: logging.Logger | None = None) -> None:
    """Convert all tables from SQLite database to Excel file"""
    if logger is None:
        logger = logging.getLogger(__name__)
    
    # Validate database file exists
    if not os.path.exists(db_path):
        raise FileNotFoundError(f"Database file not found: {db_path}")
    
    # Create output/ folder if it doesn't exist
    output_dir = os.path.dirname(output_path)
    if output_dir:
        Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    # Get list of tables
    tables = get_all_tables(db_path)
    
    if not tables:
        raise ValueError("Database does not contain any tables")
    
    logger.info(f"\nFound {len(tables)} table(s) in database:")
    for table in tables:
        logger.info(f"  - {table}")
    
    # Connect to database and create Excel writer
    with sqlite3.connect(db_path) as conn, pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for table in tables:
            # Read table into DataFrame - using quoted identifier to prevent SQL injection
            # SQLite uses double quotes for identifiers
            query = f'SELECT * FROM "{table}"'
            df = pd.read_sql_query(query, conn)
            
            # Rename data_format_X columns to readable names
            df = rename_data_format_columns(df, db_path)
            
            # Convert Unix timestamps to readable format
            df = convert_timestamps_to_readable(df)
            
            # Add row numbers for better readability
            df = add_row_numbers(df)
            
            # Save to sheet (sheet name is table name)
            # Excel has a 31 character limit for sheet names
            sheet_name = table[:31] if len(table) > 31 else table
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Format the worksheet
            worksheet = writer.sheets[sheet_name]
            format_worksheet(worksheet, df)
            
            logger.info(f"  Table '{table}': {len(df)} rows, {len(df.columns)} columns")
    
    logger.info(f"\nSuccess! Data saved to: {output_path}")
